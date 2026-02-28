import openpyxl, sqlite3
from pathlib import Path
from collections import Counter

# Check KH0483 anomaly
print('=== KH0483 Khipu sheet ===')
wb = openpyxl.load_workbook('data/kfg/KFG/KFG/KH0483.xlsx', read_only=True, data_only=True)
ws = wb['Khipu']
for row in ws.iter_rows(values_only=True):
    if row[0]: print(' ', row[0])
wb.close()

conn = sqlite3.connect('data/kfg/khipu_database.db')
cur = conn.cursor()
cur.execute("SELECT * FROM khipu_metadata WHERE kfg_id='KH0483'")
row = cur.fetchone()
cur.execute("PRAGMA table_info(khipu_metadata)")
cols = [c[1] for c in cur.fetchall()]
print()
print('=== KH0483 in DB ===')
for c, v in zip(cols, row):
    print(f'  {c}: {v!r}')

# PrimaryCord samples
print()
print('=== PrimaryCord sample (KH0001) ===')
wb = openpyxl.load_workbook('data/kfg/KFG/KFG/KH0001.xlsx', read_only=True, data_only=True)
ws = wb['PrimaryCord']
for row in ws.iter_rows(values_only=True):
    if row[0]: print(' ', row[0])
wb.close()

print()
print('=== PrimaryCord corpus-wide value distribution (all 709 files) ===')
begin_vals = Counter()
term_vals  = Counter()
twist_vals = Counter()
for f in sorted(Path('data/kfg/KFG/KFG').glob('*.xlsx')):
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb['PrimaryCord']
        for row in ws.iter_rows(values_only=True):
            if not row[0]: continue
            cell = str(row[0]).strip()
            if cell.startswith('Beginning:'):
                begin_vals[cell.split(':',1)[1].strip()] += 1
            elif cell.startswith('Termination:'):
                term_vals[cell.split(':',1)[1].strip()] += 1
            elif cell.startswith('Twist:'):
                twist_vals[cell.split(':',1)[1].strip()] += 1
        wb.close()
    except Exception:
        pass

print('Beginning values:')
for v, n in begin_vals.most_common():
    print(f'  {v!r:15} {n}')
print('Termination values:')
for v, n in term_vals.most_common():
    print(f'  {v!r:15} {n}')
print('Twist values:')
for v, n in twist_vals.most_common():
    print(f'  {v!r:15} {n}')

# Museum Description samples
print()
print('=== Museum Description samples (first 8) ===')
count = 0
for f in sorted(Path('data/kfg/KFG/KFG').glob('*.xlsx')):
    if count >= 8: break
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb['Khipu']
        for row in ws.iter_rows(values_only=True):
            if row[0] and 'Museum Description' in str(row[0]):
                val = str(row[0]).split(':', 1)
                print(f'  {f.stem}: {val[1][:70] if len(val)>1 else "(empty)"}')
                count += 1
        wb.close()
    except Exception:
        pass

# check what value the DB primary_cord table has vs missing fields
print()
print('=== DB primary_cord table columns ===')
cur.execute("PRAGMA table_info(primary_cord)")
for row in cur.fetchall():
    print(' ', row)

# check alt_value semantics - are there cords where alt_value != 0 that might be the 'real' value?
print()
print('=== alt_value != 0 sample (top 10) ===')
cur.execute("""
    SELECT kfg_id, cord_name, value, alt_value, knots, color
    FROM cords WHERE alt_value != 0 AND alt_value IS NOT NULL
    LIMIT 10
""")
for r in cur.fetchall():
    print(f'  {r[0]} {r[1]:12} value={r[2]:5} alt={r[3]:5}  knots={str(r[4])[:40]}')

# Check position column - what does it mean? 
print()
print('=== position col: sample of non-NULL rows ===')
cur.execute("""
    SELECT kfg_id, cord_name, position, group_idx, position_in_group, hierarchy_level
    FROM cords WHERE position IS NOT NULL
    LIMIT 12
""")
for r in cur.fetchall():
    print(f'  {r[0]} {r[1]:12} pos={r[2]:8} grp={r[3]} pos_in_grp={r[4]} lvl={r[5]}')

conn.close()
