"""
Full XLS audit: sheet names, Cords column coverage, CordGroups row formats,
Khipu/PrimaryCord field keys, Value column usage, and outlier-file structures.
"""
import openpyxl, re, sqlite3
from pathlib import Path
from collections import defaultdict, Counter

KFG_DIR = Path('data/kfg/KFG/KFG')
DB_PATH = Path('data/kfg/khipu_database.db')

# ── 1. Collect all sheet names across every file ──────────────────────────────
print("="*70)
print("1. SHEET NAMES ACROSS ALL FILES")
print("="*70)
sheet_name_files = defaultdict(list)
missing_expected_sheets = []

all_xlsx = sorted(KFG_DIR.glob('*.xlsx'))
print(f"Scanning {len(all_xlsx)} XLSX files...\n")

for f in all_xlsx:
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        for s in sheets:
            sheet_name_files[s].append(f.stem)
        expected = {'Khipu', 'PrimaryCord', 'Cords', 'CordGroups'}
        missing = expected - set(sheets)
        extra = set(sheets) - expected
        if missing or extra:
            missing_expected_sheets.append((f.stem, missing, extra, sheets))
    except Exception as e:
        print(f"  ERROR {f.stem}: {e}")

all_sheet_names = sorted(sheet_name_files.keys())
total_files = len(all_xlsx)
print(f"Distinct sheet names found: {len(all_sheet_names)}")
for s in all_sheet_names:
    count = len(sheet_name_files[s])
    flag = '' if count == total_files else f'  *** only in {count}/{total_files} files ***'
    print(f"  {s:25} {count:5}{flag}")

print(f"\nFiles with missing OR extra sheets: {len(missing_expected_sheets)}")
for kfg_id, missing, extra, actual in missing_expected_sheets[:30]:
    parts = []
    if missing: parts.append(f"MISSING={missing}")
    if extra:   parts.append(f"EXTRA={extra}")
    print(f"  {kfg_id:15} {', '.join(parts)}")

# ── 2. Cords sheet column names across all files ──────────────────────────────
print()
print("="*70)
print("2. CORDS SHEET COLUMN NAMES (frequency across all files)")
print("="*70)
cord_col_counter = Counter()

for f in all_xlsx:
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        if 'Cords' not in wb.sheetnames:
            wb.close()
            continue
        ws = wb['Cords']
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            cols = [str(c).strip() for c in row if c is not None]
            for c in cols:
                cord_col_counter[c] += 1
            break
        wb.close()
    except Exception:
        pass

# Columns parser currently uses
USED_COLS = {'Cord_Name','Twist','Attachment','Knots','Length','Termination',
             'Thickness','Color','Value','Alt_Value','Position','Notes'}

print(f"\nColumn name  (count / {total_files} files):")
for col, cnt in sorted(cord_col_counter.items(), key=lambda x: -x[1]):
    used = 'USED' if col in USED_COLS else 'NOT_USED'
    freq = '' if cnt >= total_files - 5 else f'  *** only in {cnt} files ***'
    print(f"  {col:30} {cnt:5}  [{used}]{freq}")

# ── 3. CordGroups sheet row format variants ───────────────────────────────────
print()
print("="*70)
print("3. CORDGROUPS ROW FORMAT VARIANTS")
print("="*70)

RE_RANGE  = re.compile(r'[\d.]+cm\s+group\s+of\s+\d+\s+pendants\s+\((\d+)-(\d+)\)', re.IGNORECASE)
RE_SINGLE = re.compile(r'[\d.]+cm\s+1\s+\((\d+)\)', re.IGNORECASE)

row_type_counter = Counter()       # 'range', 'single', 'comment', 'unmatched'
unmatched_rows   = Counter()
unmatched_examples = defaultdict(list)

for f in all_xlsx:
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        if 'CordGroups' not in wb.sheetnames:
            wb.close()
            continue
        ws = wb['CordGroups']
        for row in ws.iter_rows(values_only=True):
            cell = row[0]
            if not cell:
                continue
            cell_str = str(cell).strip()
            if cell_str.startswith('!'):
                row_type_counter['comment'] += 1
                continue
            if RE_RANGE.match(cell_str):
                row_type_counter['range'] += 1
            elif RE_SINGLE.match(cell_str):
                row_type_counter['single'] += 1
            else:
                row_type_counter['unmatched'] += 1
                key = re.sub(r'[\d.]+', 'N', cell_str)[:60]
                unmatched_rows[key] += 1
                if len(unmatched_examples[key]) < 3:
                    unmatched_examples[key].append((f.stem, cell_str))
        wb.close()
    except Exception:
        pass

print(f"Row type counts:")
for t, cnt in sorted(row_type_counter.items(), key=lambda x: -x[1]):
    print(f"  {t:15} {cnt:7,}")

print(f"\nUnrecognised patterns ({len(unmatched_rows)} distinct):")
for key, cnt in sorted(unmatched_rows.items(), key=lambda x: -x[1])[:20]:
    print(f"\n  [{cnt}x] {key}")
    for kid, ex in unmatched_examples[key][:2]:
        print(f"         {kid}: {ex!r}")

# ── 4. Khipu sheet key variants ───────────────────────────────────────────────
print()
print("="*70)
print("4. KHIPU SHEET KEY VARIANTS")
print("="*70)
KNOWN_KHIPU_KEYS = {
    'KFG_Name','Aliases','Contributors','KFG URL','Museum Name','Museum Number',
    'Museum City/State','Museum Country','Museum URL','Provenance','Region',
    'Creation_Date','Excel Write Date','Excel File Creator'
}
unknown_khipu_keys  = Counter()
unknown_khipu_ex    = defaultdict(list)

for f in all_xlsx:
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        if 'Khipu' not in wb.sheetnames:
            wb.close()
            continue
        ws = wb['Khipu']
        for row in ws.iter_rows(values_only=True):
            cell = row[0]
            if not cell: continue
            cell_str = str(cell).strip()
            if ':' in cell_str:
                key = cell_str.split(':', 1)[0].strip()
                if key not in KNOWN_KHIPU_KEYS:
                    unknown_khipu_keys[key] += 1
                    if len(unknown_khipu_ex[key]) < 2:
                        unknown_khipu_ex[key].append(f"{f.stem}: {cell_str}")
        wb.close()
    except Exception:
        pass

if unknown_khipu_keys:
    for k, cnt in unknown_khipu_keys.most_common(15):
        print(f"  {k:40} {cnt} files")
        for ex in unknown_khipu_ex[k][:1]:
            print(f"      {ex}")
else:
    print("  All Khipu sheet keys are accounted for.")

# ── 5. PrimaryCord sheet key variants ────────────────────────────────────────
print()
print("="*70)
print("5. PRIMARYCORD SHEET KEY VARIANTS")
print("="*70)
KNOWN_PC_KEYS = {'Structure','Thickness','Length','Color','Fiber'}
unknown_pc_keys = Counter()
unknown_pc_ex   = defaultdict(list)

for f in all_xlsx:
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        if 'PrimaryCord' not in wb.sheetnames:
            wb.close()
            continue
        ws = wb['PrimaryCord']
        for row in ws.iter_rows(values_only=True):
            cell = row[0]
            if not cell: continue
            cell_str = str(cell).strip()
            if ':' in cell_str:
                key = cell_str.split(':', 1)[0].strip()
                if key not in KNOWN_PC_KEYS:
                    unknown_pc_keys[key] += 1
                    if len(unknown_pc_ex[key]) < 2:
                        unknown_pc_ex[key].append(f"{f.stem}: {cell_str!r}")
        wb.close()
    except Exception:
        pass

if unknown_pc_keys:
    for k, cnt in unknown_pc_keys.most_common(15):
        print(f"  {k:40} {cnt} files")
        for ex in unknown_pc_ex[k][:1]:
            print(f"      {ex}")
else:
    print("  All PrimaryCord sheet keys are accounted for.")

# ── 6. DB column null audit ───────────────────────────────────────────────────
print()
print("="*70)
print("6. DB COLUMN NULL AUDIT (cords table)")
print("="*70)
conn = sqlite3.connect(DB_PATH)
cur  = conn.cursor()
cur.execute("SELECT COUNT(*) FROM cords")
total_db = cur.fetchone()[0]
cur.execute("PRAGMA table_info(cords)")
cols = [row[1] for row in cur.fetchall()]
for col in cols:
    cur.execute(f"SELECT COUNT(*) FROM cords WHERE [{col}] IS NOT NULL")
    nonnull = cur.fetchone()[0]
    pct = 100 * nonnull / total_db if total_db else 0
    flag = '  *** ALWAYS NULL - never populated ***' if nonnull == 0 else ''
    print(f"  {col:25} {nonnull:6}/{total_db} ({pct:5.1f}%){flag}")

# ── 7. Value-column vs knot-embedded value coverage ──────────────────────────
print()
print("="*70)
print("7. VALUE COLUMN vs KNOT-STRING EMBEDDED VALUE")
print("="*70)
cur.execute("""
    SELECT
        COUNT(*)                                                     AS total,
        SUM(CASE WHEN value IS NOT NULL     THEN 1 ELSE 0 END)       AS has_value_col,
        SUM(CASE WHEN knots IS NOT NULL     THEN 1 ELSE 0 END)       AS has_knots,
        SUM(CASE WHEN value IS NOT NULL AND knots IS NOT NULL THEN 1 ELSE 0 END) AS both
    FROM cords
""")
r = cur.fetchone()
print(f"  Total cords:                {r[0]:,}")
print(f"  value col non-NULL:         {r[1]:,}  ({100*r[1]/r[0]:.1f}%)")
print(f"  knots col non-NULL:         {r[2]:,}  ({100*r[2]/r[0]:.1f}%)")
print(f"  both value+knots set:       {r[3]:,}")

# Sample: value set in Value col but NOT embedded in knots string
cur.execute("""
    SELECT kfg_id, cord_name, value, alt_value, knots FROM cords
    WHERE value IS NOT NULL AND value != 0
      AND (knots IS NULL OR knots NOT GLOB '*,*')
    LIMIT 8
""")
rows = cur.fetchall()
print(f"\n  Cords where value comes from DB 'Value' col (NOT knot-embedded) [{len(rows)}+ examples]:")
for r in rows:
    print(f"    {r[0]} {r[1]}  value={r[2]} alt={r[3]}  knots={r[4]}")

# ── 8. Optional cord columns: termination, position, alt_value ───────────────
print()
print("="*70)
print("8. OPTIONAL CORD COLUMNS DETAIL")
print("="*70)
for col in ('termination','position','alt_value','twist','attachment','notes','length','thickness'):
    cur.execute(f"SELECT COUNT(*) FROM cords WHERE [{col}] IS NOT NULL")
    n = cur.fetchone()[0]
    pct = 100*n/total_db
    print(f"  {col:20} {n:6,} / {total_db:,}  ({pct:.1f}%)")

# alt_value distribution
cur.execute("SELECT alt_value, COUNT(*) n FROM cords WHERE alt_value IS NOT NULL GROUP BY alt_value ORDER BY n DESC LIMIT 8")
rows = cur.fetchall()
if rows:
    print(f"\n  alt_value value distribution (top 8):")
    for r in rows:
        print(f"    {r[0]:10} : {r[1]:,}")

# termination value distribution
cur.execute("SELECT termination, COUNT(*) n FROM cords WHERE termination IS NOT NULL GROUP BY termination ORDER BY n DESC LIMIT 8")
rows = cur.fetchall()
if rows:
    print(f"\n  termination value distribution (top 8):")
    for r in rows:
        print(f"    {r[0]:10} : {r[1]:,}")

# ── 9. Cord type / twist coverage ─────────────────────────────────────────────
print()
print("="*70)
print("9. TWIST VALUES (cord-ply direction)")
print("="*70)
cur.execute("SELECT twist, COUNT(*) n FROM cords GROUP BY twist ORDER BY n DESC")
for r in cur.fetchall():
    print(f"  {str(r[0]):15} {r[1]:,}")

# ── 10. Outlier / non-KH files ────────────────────────────────────────────────
print()
print("="*70)
print("10. OUTLIER FILES (non-KH: CM, HM, KT, RA, RS) + KH split variants")
print("="*70)
non_kh  = [f for f in all_xlsx if not re.match(r'KH\d{4}$', f.stem)]
split_kh= [f for f in all_xlsx if re.match(r'KH\d{4}[A-Z]', f.stem)]

for f in non_kh + split_kh:
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        sheets = wb.sheetnames
        cord_cols, cord_rows = [], 0
        if 'Cords' in sheets:
            ws = wb['Cords']
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                cord_cols = [str(c).strip() for c in row if c is not None]
                break
            cord_rows = sum(1 for _ in ws.iter_rows()) - 1
        # check CordGroups
        has_cg = 'CordGroups' in sheets
        # peek at DB
        cur.execute("SELECT COUNT(*) FROM cords WHERE kfg_id=?", (f.stem,))
        db_cords = cur.fetchone()[0]
        wb.close()
        print(f"  {f.stem:15} sheets={[s[:8] for s in sheets]}")
        print(f"               cord_cols={cord_cols}")
        print(f"               xlsx_cords={cord_rows}  db_cords={db_cords}  has_CordGroups={has_cg}")
    except Exception as e:
        print(f"  {f.stem}: ERROR {e}")

conn.close()
print()
print("AUDIT COMPLETE")
