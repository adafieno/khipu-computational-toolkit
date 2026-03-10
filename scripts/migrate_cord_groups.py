"""
Migrate cord group assignments into the KFG database.

Parses the CordGroups Excel sheet from each of the 709 khipu files and
adds group_idx / position_in_group to every pendant cord row.

cord_index [group_idx, position_in_group] is exactly what the KFG ground-truth
summation files use to reference individual cords.
"""

from pathlib import Path
import sys
import re
import sqlite3
import openpyxl
import argparse

src_path = Path(__file__).parent.parent / 'src'
sys.path.insert(0, str(src_path))
from config_kfg import get_kfg_config


# ─── parsing ──────────────────────────────────────────────────────────────────

# e.g. "13.5cm group of 46 pendants (2-47) space of 0.5cm"
RE_RANGE = re.compile(r'[\d.]+cm\s+group\s+of\s+\d+\s+pendants\s+\((\d+)-(\d+)\)', re.IGNORECASE)
# e.g. "9.0cm 1 (1)"
RE_SINGLE = re.compile(r'[\d.]+cm\s+1\s+\((\d+)\)', re.IGNORECASE)


def parse_cord_groups(excel_path: Path) -> dict:
    """
    Parse the CordGroups sheet and return:
        { pendant_num: (group_idx, position_in_group), ... }
    """
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        if 'CordGroups' not in wb.sheetnames:
            wb.close()
            return {}

        ws = wb['CordGroups']
        assignments = {}
        group_idx = 0

        for row in ws.iter_rows(values_only=True):
            cell = row[0]
            if not cell:
                continue
            cell_str = str(cell).strip()
            if cell_str.startswith('!'):       # comment line
                continue

            m_range = RE_RANGE.match(cell_str)
            if m_range:
                start, end = int(m_range.group(1)), int(m_range.group(2))
                for pos, p_num in enumerate(range(start, end + 1)):
                    assignments[p_num] = (group_idx, pos)
                group_idx += 1
                continue

            m_single = RE_SINGLE.match(cell_str)
            if m_single:
                p_num = int(m_single.group(1))
                assignments[p_num] = (group_idx, 0)
                group_idx += 1
                continue

        wb.close()
        return assignments

    except Exception as e:
        return {}


# ─── database migration ────────────────────────────────────────────────────────

def add_group_columns(conn: sqlite3.Connection):
    """Add group_idx and position_in_group columns if they don't exist."""
    cur = conn.cursor()
    existing = {row[1] for row in cur.execute('PRAGMA table_info(cords)')}
    if 'group_idx' not in existing:
        cur.execute('ALTER TABLE cords ADD COLUMN group_idx INTEGER')
    if 'position_in_group' not in existing:
        cur.execute('ALTER TABLE cords ADD COLUMN position_in_group INTEGER')
    conn.commit()
    print('✓ Columns group_idx / position_in_group ready')


def populate_groups(kfg_dir: Path, db_path: Path, limit: int = 0):
    """
    Walk every Excel file, parse its CordGroups sheet, and update cords.
    Only pendant cords (hierarchy_level = 0) receive group assignments.
    """
    excel_files = sorted(kfg_dir.glob('*.xlsx'))
    if limit:
        excel_files = excel_files[:limit]

    conn = sqlite3.connect(db_path)
    add_group_columns(conn)

    total_assigned = 0
    total_missing  = 0
    errors         = []

    print(f'Processing {len(excel_files)} Excel files...')
    print('-' * 80)

    for i, xlsx in enumerate(excel_files):
        kfg_id = xlsx.stem

        assignments = parse_cord_groups(xlsx)
        if not assignments:
            errors.append(f'{kfg_id}: no CordGroups sheet or empty')
            continue

        cur = conn.cursor()
        assigned = 0
        missing  = 0

        # Only top-level pendants (hierarchy_level=0) get group assignments
        cur.execute(
            'SELECT cord_id, pendant_num FROM cords '
            'WHERE kfg_id=? AND hierarchy_level=0',
            (kfg_id,)
        )
        for cord_id, pendant_num in cur.fetchall():
            if pendant_num in assignments:
                g, p = assignments[pendant_num]
                cur.execute(
                    'UPDATE cords SET group_idx=?, position_in_group=? WHERE cord_id=?',
                    (g, p, cord_id)
                )
                assigned += 1
            else:
                missing += 1

        conn.commit()
        total_assigned += assigned
        total_missing  += missing

        if (i + 1) % 50 == 0 or (i + 1) == len(excel_files):
            print(f'  [{i+1:4d}/{len(excel_files)}]  {kfg_id}  '
                  f'assigned={assigned}  missing={missing}')

    conn.close()

    print()
    print('='*80)
    print('MIGRATION COMPLETE')
    print('='*80)
    print(f'  Total pendants assigned: {total_assigned:,}')
    print(f'  Total pendants missing:  {total_missing:,}')
    if errors:
        print(f'  Files with issues ({len(errors)}):')
        for e in errors[:10]:
            print(f'    {e}')
    print()


def verify(db_path: Path):
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    cur.execute('SELECT COUNT(*) FROM cords WHERE hierarchy_level=0 AND group_idx IS NOT NULL')
    assigned = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM cords WHERE hierarchy_level=0 AND group_idx IS NULL')
    missing = cur.fetchone()[0]
    cur.execute('SELECT MAX(group_idx)+1, COUNT(DISTINCT group_idx) FROM cords WHERE kfg_id="KH0001"')
    kh1 = cur.fetchone()

    print('Verification:')
    print(f'  Pendants with group_idx:    {assigned:,}')
    print(f'  Pendants without group_idx: {missing:,}')
    print(f'  KH0001: {kh1[0]} groups, {kh1[1]} distinct group_idx values')

    # Show KH0001 sample
    cur.execute(
        'SELECT cord_name, pendant_num, group_idx, position_in_group '
        'FROM cords WHERE kfg_id="KH0001" AND hierarchy_level=0 '
        'ORDER BY group_idx, position_in_group LIMIT 15'
    )
    rows = cur.fetchall()
    print()
    print('  KH0001 sample (first 15):')
    print(f'  {"cord_name":10} {"pendant_num":12} {"group_idx":10} {"pos_in_group":12}')
    for r in rows:
        print(f'  {str(r[0]):10} {str(r[1]):12} {str(r[2]):10} {str(r[3]):12}')

    conn.close()


def main():
    parser = argparse.ArgumentParser(description='Populate cord group assignments in KFG database')
    parser.add_argument('--limit', type=int, default=0, help='Limit to N khipus (0=all)')
    args = parser.parse_args()

    config     = get_kfg_config()
    db_path    = config.get_database_path()
    kfg_dir    = Path('data/kfg/KFG/KFG')

    print('='*80)
    print('KFG CORD GROUP MIGRATION')
    print('='*80)
    print(f'Database: {db_path}')
    print(f'KFG dir:  {kfg_dir}')
    print()

    populate_groups(kfg_dir, db_path, limit=args.limit)
    verify(db_path)


if __name__ == '__main__':
    main()
